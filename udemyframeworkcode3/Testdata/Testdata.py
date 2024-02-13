import openpyxl
class testdata:

    test_dataforload=[{"firstname":"Raveendra","lastname":"sampathi","email":"srn456@.com",
                      "password":"12345678","day":"1","month":"2","year":"2022","gender":"Male"},
                      ]

    @staticmethod
    def gettestdata(test_case_name):

        file = "D:\Datasets\d1.xlsx"
        workbook = openpyxl.load_workbook(file)
        sheet = workbook["Sheet1"]
        Dict = {}
        rows = sheet.max_row
        columns = sheet.max_column
        for i in range(1, rows + 1):
            if sheet.cell(i, 1).value == test_case_name:  # its give the specific data

                for j in range(2, columns + 1):
                    #dict["name"]="raveendra
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
                    #print(sheet.cell(i, j).value, end=" ") not required
                #print()
        return[Dict]