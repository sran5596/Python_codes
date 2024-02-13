#write the different data in excel
import openpyxl
#file-workbook-sheet-row-column
file="D:\Datasets\ddtselenium.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet2"]
sheet.cell(1,1).value="Name"
sheet.cell(1,2).value="Age"
sheet.cell(1,3).value="Salary"
sheet.cell(2,1).value="Basha"
sheet.cell(2,2).value=23
sheet.cell(2,3).value="100crore"
sheet.cell(3,1).value="Raveendra"
sheet.cell(3,2).value=24
sheet.cell(3,3).value="120crore"
sheet.cell(4,1).value="Palapyakit"
sheet.cell(4,2).value=23
sheet.cell(4,3).value="100crore"
workbook.save(file)