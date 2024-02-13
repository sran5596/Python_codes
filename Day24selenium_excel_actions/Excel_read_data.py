import openpyxl
#strecture: file-workbook-sheet-rows-columns
file="D:\Datasets\d1.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]
#identify the max_row
rows=sheet.max_row # its count the rows
cols=sheet.max_column # its count the columns
for r in range(1,rows+1): # here used the range(n-1) it take so used the rows+1 it read last row
    for c in range(1,cols+1): #cols+1 it read the last one also
        print(sheet.cell(r,c).value,end="       ")
    print() # it gives the gap between row and row


