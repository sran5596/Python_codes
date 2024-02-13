#read data
def read():
    import openpyxl
    file="D:\Datasets\d1.xlsx" #file-workbook-sheet-rows-columns
    workbook=openpyxl.load_workbook(file)
    sheet=workbook["Sheet1"]
    rows=sheet.max_row
    columns=sheet.max_column
    for r in range(1,rows+1):
        #if sheet.cell(i, 1).value == "Name":  # its print the particular data related only
        for c in range(1,columns+1):
            print(sheet.cell(r,c).value,end="       ")
        print()
#same data write
def write():
    import openpyxl
    file="D:\Datasets\ddtselenium.xlsx"
    workbook=openpyxl.load_workbook(file)
    sheet=workbook["Sheet1"]
    for r in range(1,6):

        for c in range(1,6):
            sheet.cell(r,c).value="Raveendra"
    workbook.save(file)
#write the different data
def write_data():
    import openpyxl
    file="D:\Datasets\ddtselenium.xlsx"
    workbook=openpyxl.load_workbook(file)
    sheet=workbook["Sheet4"]
    #its different data store process
    sheet.cell(1,1).value="Name"
    sheet.cell(1,2).value="age"
    sheet.cell(1,3).value="property"
    sheet.cell(2,1).value="Raveendra"
    sheet.cell(2,2).value=25
    sheet.cell(2,3).value="120crore"
    workbook.save(file)
    rows1=sheet.max_row
    columns1=sheet.max_column
    for r in range(1,rows1+1):
        for c in range(1,columns1+1):
            print(sheet.cell(r,c).value,end="   ")
        print()

read()
write_data()

# #specific data relatd to print
# import openpyxl
# file="path"
# worksheet=openpyxl.load_workbook(file)
# sheet=worksheet["Sheet1"]
# rows=sheet.max_row
# columns=sheet.max_column
# for r in range(1,rows+1):
#     #if condition prints only the needed data only
#     if sheet.cell(r,1).value=="name":
#
#     for c in range(2,columns+1):
#         print(sheet.cell(r,c).value,end="       ")
#     print()

