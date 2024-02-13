import openpyxl

file="D:\Datasets\d1.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]
Dict={}
rows=sheet.max_row
columns=sheet.max_column
for i in range(1,rows+1):
    if sheet.cell(i,1).value=="Divya": #its give the specific data

        for j in range(2,columns+1):
            Dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i, column=j).value
            print(sheet.cell(i,j).value,end=" ")
        print()
print(Dict)