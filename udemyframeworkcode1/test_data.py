import pytest
import openpyxl
class datasets: #its speciallly for the datausage purpose

    test_data_sets=[{"firstname":"Raveendra","lastname":"sampathi",},
                    {"firstname":"divya","lastname":"sampathi"}]

#below code use where you need in the dataload page
# @pytest.fixture(params=[{"firstname":"Raveendra","lastname":"sampathi",},
#                     {"firstname":"divya","lastname":"sampathi"}])
#@pytest.fixture(params=datasets.test_data_sets)
# def test_data_sets(request):
#     return request.param
    @staticmethod
    def test_data(test_case_name):
        @staticmethod
        def gettestdata(test_case_name):

            file = "D:\Datasets\d1.xlsx" # here pass the path of excel remain all same
            workbook = openpyxl.load_workbook(file)
            sheet = workbook["Sheet1"]
            Dict = {}
            rows = sheet.max_row
            columns = sheet.max_column
            for i in range(1, rows + 1):
                if sheet.cell(i, 1).value == test_case_name:  # its give the specific data

                    for j in range(2, columns + 1):
                        # dict["name"]="raveendra
                        Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
                        # print(sheet.cell(i, j).value, end=" ") not required
                    # print()
            return [Dict]