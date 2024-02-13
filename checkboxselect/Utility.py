import openpyxl
from openpyxl.styles import PatternFill

#get total rows count
def getRowCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return(sheet.max_row)


#get total columns count
def getColumnCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return(sheet.max_column)

#read the data form excel
def readData(file,sheetName,rownum,columnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return sheet.cell(rownum,columnno).value

#write the data in excel
def writeData(file,sheetName,rownum,columnno,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    sheet.cell(rownum,columnno).value=data
    workbook.save(file)

#Fill the colours to cells
def fillGreenColour(file,sheetName,rownum,columnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    greenFill=PatternFill(start_color='60b212',
                          end_color='60b212',
                          fill_type='solid')
    sheet.cell(rownum,columnno).fill=greenFill
    workbook.save(file)

#fill the red colur
def fillRedColour(file,sheetName,rownum,columnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    redFill=PatternFill(start_color='ff0000',
                        end_color='ff0000',
                        fill_type='solid')
    sheet.cell(rownum,columnno).fill=redFill
    workbook.save(file)


