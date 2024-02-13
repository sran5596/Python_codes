import openpyxl

from Day24selenium_excel_actions import Utility
#read excel
# file="D:\Datasets\d1.xlsx"
# book=openpyxl.load_workbook(file)
# sheet=book["Sheet1"]
# # columns=sheet.max_column
# # rows=sheet.max_row
# # print(columns)
# # print(rows)
# # for r in range(1,rows+1):
# #     for c in range(1,columns+1):
# #         print(sheet.cell(r,c).value,end="        ")
# # print()
#write excel
# file="D:\Datasets\write.xlsx"
# book=openpyxl.load_workbook(file)
# sheet=book['Sheet1']
# sheet.cell(1,1).value="Name"
# sheet.cell(1,2).value="Email"
# sheet.cell(2,1).value="rahul"
# sheet.cell(2,2).value="kumar"
#
# book.save(file)
# from checkboxselect.Utility import *
# from selenium import webdriver
# import time
# from Day24selenium_excel_actions import Utility
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.select import Select
# from selenium.webdriver.edge.service import Service
# from selenium.webdriver.support import expected_conditions as EC
# from selenium.webdriver.support.wait import WebDriverWait
# obj=Service("D:/driveers/msedgedriver.exe")
# driver=webdriver.Edge(service=obj)
# mywait=WebDriverWait(driver,20)
# #below code for read the adata
# file="D:/Datasets/fbdata.xlsx"
# rows=Utility.getRowCount(file,"Sheet1")
# print(rows)
# columns=Utility.getColumnCount(file,"Sheet1")
# print(columns)
#
# for r in range(2,rows+1):
#     fname=Utility.readData(file,"Sheet1",r,1)
#     lname=Utility.readData(file,"Sheet1",r,2)
#     Email=Utility.readData(file,"Sheet1",r,3)
#     pws=Utility.readData(file,"Sheet1",r,4)
#     day=Utility.readData(file,"Sheet1",r,5)
#     month=Utility.readData(file,"Sheet1",r,6)
#     year=Utility.readData(file,"Sheet1",r,7)
#     gender=Utility.readData(file,"Sheet1",r,8)
# #here application code
#     driver.get("https://www.facebook.com/signup")
#     time.sleep(2)
#     f1=driver.find_element(By.XPATH,"//input[@class='inputtext _58mg _5dba _2ph-' and @id='u_0_b_rP']")
#     # mywait.until(EC.presence_of_element_located((By.CSS_SELECTOR,"input#u_0_b_rP"))).send_keys(fname)
#     f1.send_keys(fname)
#     time.sleep(2)
#     driver.find_element(By.XPATH, "//input[@id='u_0_d_d1']").send_keys(lname)
#     driver.find_element(By.XPATH, "//input[@id='u_0_g_xz']").send_keys(Email)
#     driver.find_element(By.XPATH, "//input[@id='password_step_input']").send_keys(pws)
#     dayslt=Select(driver.find_element(By.XPATH, "//input[@id='day']"))
#     dayslt.select_by_value(day)
#     monthslt=Select(driver.find_element(By.XPATH, "//input[@id='month']"))
#     monthslt.select_by_value(month)
#     yearslt=Select(driver.find_element(By.XPATH, "//input[@id='year']"))
#     yearslt.select_by_value(year)
#     gen1=driver.find_elements(By.XPATH,"//span[@class='_5k_2 _5dba']")
#     if gen1.text=='gender':
#         gen1.click()
# time.sleep(5)


# #MySql data
# import mysql.connector
# con=mysql.connector.connect(host="localhost",port=3306,user="root",password="sraveendra009403@@",database="p1")
# curs=con.cursor()
# # curs.execute("insert into d1 values(1,'Raveendra')")
# curs.execute("select * from d1")
# for i in curs:
#     print(i[0])
#     print(i[1])
# con.commit()
# con.close()




