import openpyxl
file="D:\Datasets\write.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]
sheet.cell(1,1).value="Name"
sheet.cell(1,2).value="Age"
sheet.cell(2,1).value="Raveendra"
sheet.cell(2,2).value=25
workbook.save(file)
