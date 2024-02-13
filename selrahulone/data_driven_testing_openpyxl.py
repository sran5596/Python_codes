from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import openpyxl
file="D:\Datasets\openpyxl_operations.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]
totalrows=sheet.max_row
totalcols=sheet.max_column
for r in range(1,totalrows+1):
    for c in range(1,totalcols+1):
        print(sheet.cell(r,c).value,end="")

    print()


